#!/usr/bin/env python3
"""
Merge all battery CSV files with fixed metadata from XLSX into one big CSV.

Each row = one cycle of one battery, with cycle-varying data +
fixed battery metadata joined by Cell Name.
"""

import csv
import os
import sys
from pathlib import Path

import openpyxl

# Config
DATA_DIR = Path(__file__).parent / "Summary"
METADATA_FILE = Path(__file__).parent / "Summary_cells_MC_YB_03-06-2026.xlsx"
OUTPUT_FILE = Path(__file__).parent / "all_batteries_combined.csv"
CSV_ENCODING = "latin-1"
OUTPUT_ENCODING = "utf-8"


def read_metadata(path: Path) -> dict:
    """Read XLSX, return {cell_name: row_tuple}."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Cycled Cells"]
    metadata = {}
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
        if not row:
            continue
        cell_name = str(row[2]).strip() if row[2] else ""
        if cell_name:
            metadata[cell_name] = row
    return metadata


def find_csv_for_cell(cell_name: str, csv_files: list) -> str | None:
    """Find CSV file whose name starts with cell_name."""
    prefix = cell_name + "_"
    for f in csv_files:
        if f.startswith(prefix):
            return f
    return None


def main():
    # ── 1. Read metadata ──
    print(f"Reading metadata from {METADATA_FILE}...")
    metadata = read_metadata(METADATA_FILE)
    print(f"  Found {len(metadata)} batteries in XLSX.")

    # ── 2. List CSV files ──
    all_csv = sorted(
        f for f in os.listdir(DATA_DIR) if f.endswith(".csv")
    )
    print(f"  Found {len(all_csv)} CSV files in {DATA_DIR}.")

    # ── 3. Header construction ──
    cycle_headers = [
        "Cycle",
        "Cycle_Time (h)",
        "Avg_Discharge_Voltage (V)",
        "Avg_Charge_Voltage (V)",
        "Avg_Charge_Current (A)",
        "Discharge_Capacity (mAh)",
        "Charge_Capacity (mAh)",
        "Coulomb_Efficiency (%)",
        "SOH_Energy (%)",
        "Polarization (Ohm cm²)",
    ]

    meta_headers = [
        "Project",
        "Cell_Name",
        "Chemistry",
        "Cut-off",
        "Temperature",
        "C-rate",
        "Cathode_Area (cm²)",
        "CAM_Loading (mg/cm²)",
        "CAM_Mass (g)",
        "CAM_Theoretical_Capacity (mAh/g)",
        "CAM_Capacity_1st_cycle (mAh)",
        "Normalized_CAM_Capacity_2nd_cycle (mAh/g)",
        "Completed_cycles",
        "Cycles_>_80%_capacity",
        "Cycles_>_50%_capacity",
    ]

    output_headers = cycle_headers + meta_headers

    # ── 4. Write output ──
    written = 0
    matched = 0
    unmatched = []

    with open(OUTPUT_FILE, "w", encoding=OUTPUT_ENCODING, newline="") as fout:
        writer = csv.writer(fout)
        writer.writerow(output_headers)

        for fname in all_csv:
            cell_name = fname.split("_")[0]

            meta_row = metadata.get(cell_name)
            if meta_row is None:
                unmatched.append(fname)
                continue

            matched += 1

            # Build fixed metadata columns (same for all cycles of this battery)
            # meta_row indices 0..19 correspond to cols A..T in the sheet
            #   meta_row[0]  = col A (always empty)
            #   meta_row[1]  = Project
            #   meta_row[2]  = Cell Name
            #   ...
            #   meta_row[15] = Cycles > 50% capacity
            #   meta_row[16:]= unused empty columns
            # We take indices 1..15 inclusive (15 columns) to match meta_headers.
            fixed_cols = [
                str(v) if v is not None else ""
                for v in meta_row[1:16]
            ]

            filepath = DATA_DIR / fname
            with open(filepath, encoding=CSV_ENCODING, newline="") as fin:
                reader = csv.reader(fin)
                try:
                    next(reader)  # skip header row 1
                    next(reader)  # skip units row 2
                except StopIteration:
                    continue

                for data_row in reader:
                    if not data_row or all(v.strip() == "" for v in data_row):
                        continue
                    writer.writerow(data_row + fixed_cols)
                    written += 1

    # ── 5. Report ──
    print(f"\nDone. Matched {matched}/{len(all_csv)} batteries.")
    if unmatched:
        print(f"WARNING: {len(unmatched)} CSV files had no XLSX match:")
        for f in unmatched:
            print(f"  {f}")
    print(f"Wrote {written} data rows to {OUTPUT_FILE}")
    file_size = OUTPUT_FILE.stat().st_size
    print(f"  File size: {file_size / 1024 / 1024:.1f} MB")

    return 0 if not unmatched else 1


if __name__ == "__main__":
    sys.exit(main())
